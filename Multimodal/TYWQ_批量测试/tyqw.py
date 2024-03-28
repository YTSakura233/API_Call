"""
通义千问VL模型自动化测试（目前仅支持.xlsx文件）
@author YTSakura
"""
import argparse
import os
import shutil
import time
import openpyxl
import xlrd

from tyqwVL import simple_multimodal_conversation_call

date = time.strftime("%Y%m%d")


def read_excel(excel_file):
    """
    读取Excel文件内容
    :param excel_file: 文件路径
    :return:
    """
    book = ''
    if excel_file.endswith('xls'):
        book = xlrd.open_workbook(excel_file)
    elif excel_file.endswith('xlsx'):
        book = openpyxl.load_workbook(excel_file)
    return book


def getsheetname(excel_file):
    """
    获取页签名称
    :param excel_file:
    :return:
    """
    sheetname = ''
    book = read_excel(excel_file)
    if excel_file.endswith('xls'):
        sheetname = book.sheet_by_name()
    elif excel_file.endswith('xlsx'):
        sheetname = book.sheetnames
    return sheetname, book


def get_value(sheetname, book):
    sheet = book[sheetname[0]]
    nrows = sheet.max_row
    ncols = sheet.max_column
    if sheet.cell(1, 1).value != 'question' or sheet.cell(1, 2).value != 'pic' or sheet.cell(1, 3).value != 'time':
        print('表格标题错误，请确认。')
    for i in range(2, nrows + 1):
        for j in range(1, ncols + 1):
            value = sheet.cell(i, j).value
            if j == 1:
                question.append(value)
            elif j == 2:
                pic.append(value)
            else:
                times.append(value)


if __name__ == '__main__':
    isExists = os.path.exists('./history')
    if not isExists:
        os.makedirs('./history')
    if os.path.exists(f'{date}-report.txt'):
        shutil.move(f'{date}-report.txt', './history')
    question = []
    pic = []
    times = []
    result = []
    parser = argparse.ArgumentParser(description='tyqw')
    parser.add_argument('--file', type=str, help='excel file path')
    args = parser.parse_args()
    file = args.file
    sheetname, book = getsheetname(file)
    get_value(sheetname, book)

    for i in range(len(question)):
        result.append(f"'{question[i]}'的{times[i]}次测试结果：")
        for j in range(times[i]):
            test_result = simple_multimodal_conversation_call(pic[i], question[i])
            result.append(test_result)

    report_file = open(f'{date}-report.txt', 'w', encoding='utf-8')
    for item in result:
        report_file.write(str(item) + '\n')